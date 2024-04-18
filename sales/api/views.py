from django.shortcuts import render
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from sales.models import *
from .serializers import CustomerSerializer, SaleSerializer, SalesReturnSerializer
from django.db import transaction
from datetime import date, timedelta
from django.db.models import Sum
from django.db.models import F
from core.models import *
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from django.utils.encoding import smart_str
from rest_framework.authentication import TokenAuthentication

class SaleCreateView(APIView):
    def get(self, request, format=None):
        sales = Sale.objects.filter(sale_return=False).order_by('-id')
        serializer = SaleSerializer(sales, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        customer_data = request.data.get('customer')
        saleitems_data = request.data.get('saleitems')

        customer, created = Customer.objects.get_or_create(**customer_data)

        with transaction.atomic():
            sale_data = {
                'customer': customer,
                # 'vat_percentage': request.data.get('vat_percentage'),
                # 'tax_percentage': request.data.get('tax_percentage'),
                'discount_percentage': Decimal(request.data.get('discount_percentage')),
                'subtotal': Decimal(request.data.get('subtotal')),
                'total': Decimal(request.data.get('total')),
                'delivery_cost': Decimal(request.data.get('delivery_cost')),
                'curierInvoice': request.data.get('curierInvoice')
            }
            sale = Sale.objects.create(**sale_data)

            for item_data in saleitems_data:
                inventory_item = InventoryItem.objects.get(id=int(item_data['id']))

                if inventory_item is None:
                    raise ValueError("The field 'unit' is likely intended for available quantity. Rename it to 'quantity' or 'available_quantity'.")

                new_quantity = inventory_item.unit - item_data['quantity']
                if new_quantity >= 0:
                    inventory_item.unit = new_quantity
                    inventory_item.save()
                else:
                    raise ValueError("Insufficient stock for sale. Item ID: " + str(inventory_item.id))

                SaleItem.objects.create(
                    sale=sale,
                    item=inventory_item,
                    quantity=item_data['quantity'],
                    size=item_data.get('size')
                )

        serializer = SaleSerializer(sale)
        return Response(serializer.data, status=status.HTTP_201_CREATED)



class SaleRetrieveView(APIView):
    def get(self, request, pk):
        try:
            print(pk)
            sale = Sale.objects.get(pk=pk)
            serializer = SaleSerializer(sale)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Sale.DoesNotExist:
            return Response({'error': 'Sale not found'}, status=status.HTTP_404_NOT_FOUND)
        
class SalesReturnView(APIView):
    def get(self, request, format=None):
        sale_returns = SalesReturn.objects.all().order_by('-id')
        serializer = SalesReturnSerializer(sale_returns, many=True)
        return  Response(serializer.data, status=status.HTTP_200_OK)
    
    def post(self, request):
        sale_id = request.data.get('sale_id', None)
        reason = request.data.get('reason', None)

        try:
            sale = Sale.objects.get(id=sale_id)
        except Sale.DoesNotExist:
            return Response({'error': 'Invalid sale ID provided'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            sales_return = SalesReturn.objects.create(sale=sale, reason=reason)
            sale.sale_return = True
            sale.save()

            for sale_item in sale.saleitem_set.all():
                inventory_item = InventoryItem.objects.get(id=sale_item.item_id)
                inventory_item.unit += sale_item.quantity
                inventory_item.save()  # Update inventory

            serializer = SalesReturnSerializer(sales_return)
            return Response(serializer.data, status=status.HTTP_201_CREATED)



class DashboardView(APIView):
    def get(self, request):
        today = date.today()
        yesterday = today - timedelta(days=1)
        start_of_week = today - timedelta(days=today.weekday())

        # Total unit sold data
        total_unit_sold_today = Sale.objects.filter(created_date=today).aggregate(total_units_sold=Sum('saleitem_set__quantity'))['total_units_sold'] or 0
        total_unit_sold_week = Sale.objects.filter(created_date__gte=start_of_week, created_date__lte=today).aggregate(total_units_sold=Sum('saleitem_set__quantity'))['total_units_sold'] or 0
        total_unit_sold_month = Sale.objects.filter(created_date__gte=today.replace(day=1), created_date__lte=today).aggregate(total_units_sold=Sum('saleitem_set__quantity'))['total_units_sold'] or 0
        total_unit_sold_yesterday = Sale.objects.filter(created_date=yesterday).aggregate(total_units_sold=Sum('saleitem_set__quantity'))['total_units_sold'] or 0

        # Total amount sold data (using the same filtering approach)
        total_amount_sold_today = Sale.objects.filter(created_date=today).aggregate(total_sold=Sum('total'))['total_sold'] or 0
        total_amount_sold_week = Sale.objects.filter(created_date__gte=start_of_week, created_date__lte=today).aggregate(total_sold=Sum('total'))['total_sold'] or 0
        total_amount_sold_month = Sale.objects.filter(created_date__gte=today.replace(day=1), created_date__lte=today).aggregate(total_sold=Sum('total'))['total_sold'] or 0
        total_amount_sold_yesterday = Sale.objects.filter(created_date=yesterday).aggregate(total_sold=Sum('total'))['total_sold'] or 0

        # all the data for Sales return 
        today_returns = SalesReturn.objects.filter(return_date=today).count()
        week_returns = SalesReturn.objects.filter(return_date__gte=start_of_week, return_date__lte=today).count()
        month_returns = SalesReturn.objects.filter(return_date__gte=today.replace(day=1), return_date__lte=today).count()
        yesterday_returns = SalesReturn.objects.filter(return_date=yesterday).count()


        total_sales_return = SalesReturn.objects.all().count()
        unit_alerts = InventoryItem.objects.filter(unit__lt=10).values('id', 'itemName', 'unit')

        response_data = {
            'total_unit_sold': {
                'today': total_unit_sold_today,
                'week': total_unit_sold_week,
                'month': total_unit_sold_month,
                'yesterday': total_unit_sold_yesterday,
            },
            'total_amount_sold': {
                'today': total_amount_sold_today,
                'week': total_amount_sold_week,
                'month': total_amount_sold_month,
                'yesterday': total_amount_sold_yesterday,
            },
            'total_sales_return': {
                'today': today_returns,
                'week': week_returns,
                'month': month_returns,
                'yesterday': yesterday_returns,
            },
            'unit_alerts': list(unit_alerts),
        }

        return Response(response_data, status=status.HTTP_200_OK)


def get_sales_data(start_date, end_date):
    sales = Sale.objects.filter(created_date__gte=start_date, created_date__lte=end_date)
    return sales


class SalesDataView(APIView):
    def get(self, request, start_date, end_date):
        sales_data = Sale.objects.filter(created_date__gte=start_date, created_date__lte=end_date)
        serializer = SaleSerializer(sales_data)
        return Response(serializer.data, status=status.HTTP_200_OK)

def generate_excel_response(sales_data):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1).value = "Sale ID"
    worksheet.cell(row=1, column=2).value = "Customer Name"
    worksheet.cell(row=1, column=3).value = "Total"
    worksheet.cell(row=1, column=4).value = "Created Date"

    row_index = 2
    for sale in sales_data:
        worksheet.cell(row=row_index, column=1).value = str(sale.id).encode('latin1')
        worksheet.cell(row=row_index, column=2).value = str(sale.customer.name).encode('latin1')
        worksheet.cell(row=row_index, column=3).value = str(sale.total).encode('latin1')
        worksheet.cell(row=row_index, column=4).value = str(sale.created_date).encode('latin1')
        row_index += 1

    excel_data = BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)

    return excel_data


class SalesReportView(APIView):
    def get(self, request, start_date, end_date):
        try:
            start_date = date.fromisoformat(start_date)
            end_date = date.fromisoformat(end_date)
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

        sales_data = get_sales_data(start_date, end_date)

        try:
            excel_data = generate_excel_response(sales_data)
        except UnicodeDecodeError as e:
            return Response({'error': 'Unicode error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        print('successfully passed the function calling')

        filename = f'sales_report_{start_date.strftime("%Y-%m-%d")}_{end_date.strftime("%Y-%m-%d")}.xlsx'
        response = Response(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        print('success on reponse gathering')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

